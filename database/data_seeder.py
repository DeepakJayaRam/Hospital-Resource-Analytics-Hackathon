from openpyxl import Workbook

# Create workbook
wb = Workbook()

# -------- Sheet 1: Departments --------
ws_dept = wb.active
ws_dept.title = "Departments"
ws_dept.append(["ID", "Name", "Branch", "Total Beds"])
ws_dept_data = [
    [1, "Cardiology", "Main Block", 20],
    [2, "Neurology", "Main Block", 15],
    [3, "Orthopedics", "East Wing", 25],
    [4, "General Medicine", "Main Block", 30],
    [5, "Pediatrics", "North Wing", 18],
    [6, "Emergency", "Emergency Block", 40],
    [7, "ICU", "Critical Care Block", 12],
    [8, "Gynecology", "South Wing", 22],
]
for r in ws_dept_data:
    ws_dept.append(r)

# -------- Sheet 2: Doctors --------
ws_doc = wb.create_sheet("Doctors")
ws_doc.append(["ID", "Name", "Department ID", "Specialty"])
ws_doc_data = [
    [101, "Dr. Priya Sharma", 1, "Cardiac Surgeon"],
    [102, "Dr. Amit Verma", 1, "Cardiologist"],
    [103, "Dr. Neha Rao", 2, "Neurologist"],
    [104, "Dr. Rahul Mehta", 3, "Orthopedic Surgeon"],
    [105, "Dr. Ananya Iyer", 4, "General Physician"],
    [106, "Dr. Suresh Kumar", 5, "Pediatrician"],
    [107, "Dr. Kavita Nair", 6, "Emergency Medicine"],
    [108, "Dr. Arjun Singh", 7, "Intensivist"],
    [109, "Dr. Pooja Desai", 8, "Gynecologist"],
]
for r in ws_doc_data:
    ws_doc.append(r)

# -------- Sheet 3: Patients --------
ws_pat = wb.create_sheet("Patients")
ws_pat.append(["ID", "Name", "Age", "Gender", "City"])
ws_pat_data = [
    [501, "Rahul Verma", 34, "M", "Mumbai"],
    [502, "Anita Singh", 28, "F", "Delhi"],
    [503, "Karthik Reddy", 45, "M", "Hyderabad"],
    [504, "Meena Iyer", 52, "F", "Chennai"],
    [505, "Suresh Naidu", 60, "M", "Visakhapatnam"],
    [506, "Pooja Patel", 31, "F", "Ahmedabad"],
    [507, "Rohan Malhotra", 40, "M", "Bangalore"],
    [508, "Lakshmi Devi", 65, "F", "Vijayawada"],
]
for r in ws_pat_data:
    ws_pat.append(r)

# -------- Sheet 4: Admissions --------
ws_adm = wb.create_sheet("Admissions")
ws_adm.append([
    "ID", "Patient ID", "Doctor ID", "Dept ID",
    "Admission Date", "Discharge Date", "Type", "Billing Amount"
])
ws_adm_data = [
    [9001, 501, 101, 1, "2023-10-05", "", "Emergency", 15000],
    [9002, 502, 105, 4, "2023-10-06", "2023-10-08", "Elective", 8000],
    [9003, 503, 104, 3, "2023-10-07", "2023-10-12", "Elective", 22000],
    [9004, 504, 103, 2, "2023-10-09", "", "Emergency", 18000],
    [9005, 505, 108, 7, "2023-10-10", "2023-10-15", "Emergency", 30000],
    [9006, 506, 109, 8, "2023-10-11", "2023-10-13", "Elective", 12000],
]
for r in ws_adm_data:
    ws_adm.append(r)

# Save file
file_path = "Hospital_Dashboard_Data.xlsx"
wb.save(file_path)

file_path
